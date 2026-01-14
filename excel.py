# -*- coding: utf-8 -*-

import os
import re
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook

DOSYA_ADI = "adaylar.xlsx"
SAYFA_ADI = "Adaylar"

BASLIKLAR = [

#11 tane baslik ekle






]

FIELDS = [
    {"key": "Ad Soyad", "label": "Ad Soyad", "type": "entry", "hint": "Örn: Ahmet Yılmaz"}, #buradaki veriler değişebilir 
    {"key": "Başvurduğu Pozisyon", "label": "Başvurduğu Pozisyon", "type": "entry", "hint": "Örn: Elektrik Elektronik"}, #buradaki veriler değişebilir 
    {"key": "Üniversite", "label": "Üniversite", "type": "entry", "hint": "Örn: İstanbul Üniversitesi"}, #buradaki veriler değişebilir 
    {"key": "Okuduğu Bölüm", "label": "Okuduğu Bölüm", "type": "entry", "hint": "Örn: Elektrik Elektronik Mühendisliği"}, #buradaki veriler değişebilir 
    {
        "key": "Eğitim Düzeyi",
        "label": "Eğitim Düzeyi",
        "type": "combo",
        "values": ["Lise", "Önlisans", "Lisans", "Yüksek Lisans", "Doktora", "Diğer"],
        "hint": "Seçiniz",
    },
    {"key": "Mezuniyet Tarihi", "label": "Mezuniyet Tarihi", "type": "entry", "hint": "Örn: 06.2024 veya 2024-06 (boş bırakılabilir)"},     #buradaki veriler değişebilir 
    {"key": "Not Ortalaması", "label": "Not Ortalaması", "type": "entry", "hint": "Örn: 3.12/4.00 veya 78/100 (boş bırakılabilir)"},  #buradaki veriler değişebilir 
    {"key": "Doğum Yılı", "label": "Doğum Yılı", "type": "entry", "hint": "Sadece yıl: 1990 (boş bırakılabilir)"}, #buradaki veriler değişebilir 
    {"key": "İş Deneyimi Var mı?", "label": "İş Deneyimi Var mı?", "type": "radio", "values": ["Evet", "Hayır"], "hint": ""}, #buradaki veriler değişebilir 
    {"key": "Şu an Çalışıyor mu?", "label": "Şu an Çalışıyor mu?", "type": "radio", "values": ["Evet", "Hayır"], "hint": ""}, #buradaki veriler değişebilir 
    {"key": "Mevcut Firma", "label": "Mevcut Firma", "type": "entry", "hint": "Şu an çalışıyorsa yaz (boş bırakılabilir)"}, #buradaki veriler değişebilir 
]

def turkce_title(text: str) -> str:
    if text is None:
        return ""
    t = text.strip()
    if not t:
        return ""
    lower_map = str.maketrans({"I": "ı", "İ": "i"})
    t = t.translate(lower_map).lower()

    words = t.split()
    out = []
    for w in words:
        first = w[0]
        rest = w[1:]
        if first == "i":
            first_up = "İ"
        elif first == "ı":
            first_up = "I"
        else:
            first_up = first.upper()
        out.append(first_up + rest)
    return " ".join(out)

def normalize_yil(text: str) -> str:
    if text is None:
        return ""
    t = text.strip()
    if not t:
        return ""
    m = re.search(r"\b(19\d{2}|20\d{2})\b", t)  # 1900-2099
    return m.group(1) if m else ""

def excel_hazirla(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        if SAYFA_ADI in wb.sheetnames:
            ws = wb[SAYFA_ADI]
        else:
            ws = wb.create_sheet(SAYFA_ADI)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SAYFA_ADI

    # Başlık satırı yoksa oluştur
    if ws.max_row < 1 or any(ws.cell(row=1, column=i + 1).value is None for i in range(len(BASLIKLAR))):
        for i, h in enumerate(BASLIKLAR, start=1):
            ws.cell(row=1, column=i).value = h
        wb.save(path)
    return wb, ws

class App(ttk.Frame):
    def __init__(self, root):
        super().__init__(root, padding=14)
        self.root = root
        self.pack(fill="both", expand=True)

        self.wb, self.ws = excel_hazirla(DOSYA_ADI)
        self.current_row = self.ws.max_row + 1

        self.idx = 0
        self.answers = {}

        self._style()
        self._build_ui()
        self._show_field(0)

        self.root.bind("<Return>", self._on_enter)
        self.root.bind("<Escape>", lambda e: self._finish())

    def _style(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Header.TLabel", font=("Segoe UI", 15, "bold"))
        style.configure("Sub.TLabel", font=("Segoe UI", 10))
        style.configure("Field.TLabel", font=("Segoe UI", 11, "bold"))
        style.configure("Hint.TLabel", font=("Segoe UI", 9))
        style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"))
        style.configure("TButton", font=("Segoe UI", 10))
        style.configure("TEntry", font=("Segoe UI", 11))
        style.configure("TCombobox", font=("Segoe UI", 11))

    def _build_ui(self):
        header = ttk.Frame(self)
        header.pack(fill="x")
        ttk.Label(header, text="Tablo Bilgisi Girişi", style="Header.TLabel").pack(anchor="w")
        ttk.Label(header, text=f"Excel: {DOSYA_ADI}", style="Sub.TLabel").pack(anchor="w", pady=(2, 0))

        prog = ttk.Frame(self)
        prog.pack(fill="x", pady=(12, 6))
        self.progress_text = ttk.Label(prog, text="", style="Sub.TLabel")
        self.progress_text.pack(side="left")
        self.progress = ttk.Progressbar(prog, maximum=len(FIELDS), length=260)
        self.progress.pack(side="right")

        self.card = ttk.Frame(self, padding=14)
        self.card.pack(fill="both", expand=True, pady=(6, 10))

        self.field_label = ttk.Label(self.card, text="", style="Field.TLabel")
        self.field_label.pack(anchor="w")

        self.hint_label = ttk.Label(self.card, text="", style="Hint.TLabel")
        self.hint_label.pack(anchor="w", pady=(4, 10))

        self.input_container = ttk.Frame(self.card)
        self.input_container.pack(fill="x")

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", pady=(4, 0))

        self.btn_back = ttk.Button(bottom, text="⬅ Geri", command=self._back)
        self.btn_back.pack(side="left")

        self.btn_skip = ttk.Button(bottom, text="Boş Geç", command=self._skip)
        self.btn_skip.pack(side="left", padx=6)

        self.btn_next = ttk.Button(bottom, text="Sonraki  ⏎", style="Primary.TButton", command=self._next)
        self.btn_next.pack(side="left", padx=6)

        self.btn_new = ttk.Button(bottom, text="Adayı Bitir (Yeni Aday)", command=self._force_finish_candidate)
        self.btn_new.pack(side="left", padx=6)

        self.btn_exit = ttk.Button(bottom, text="Bitir & Çık (Esc)", command=self._finish)
        self.btn_exit.pack(side="right")

        ttk.Label(self, text="Kısayollar: Enter=Sonraki • Esc=Çık", style="Hint.TLabel").pack(anchor="w")

    def _clear_input_container(self):
        for w in self.input_container.winfo_children():
            w.destroy()

    def _prefill_value(self, key: str):
        """Geri gelince önceki cevabı input'a geri bas."""
        val = self.answers.get(key, "")
        self.var.set(val)

    def _show_field(self, idx: int):
        self.idx = idx
        f = FIELDS[idx]

        self.progress["value"] = idx
        self.progress_text.config(text=f"Alan {idx + 1} / {len(FIELDS)}")

        self.field_label.config(text=f["label"])
        self.hint_label.config(text=f.get("hint", ""))

        self._clear_input_container()
        self.var = tk.StringVar()

        if f["type"] == "entry":
            self.entry = ttk.Entry(self.input_container, textvariable=self.var)
            self.entry.pack(fill="x")
            self._prefill_value(f["key"])
            self.entry.focus_set()

        elif f["type"] == "combo":
            self.combo = ttk.Combobox(self.input_container, textvariable=self.var, values=f["values"], state="readonly")
            self.combo.pack(fill="x")
            self._prefill_value(f["key"])
            self.combo.focus_set()

        elif f["type"] == "radio":
            row = ttk.Frame(self.input_container)
            row.pack(anchor="w")
            # mevcut seçim geri gelsin
            self._prefill_value(f["key"])
            for val in f["values"]:
                ttk.Radiobutton(row, text=val, value=val, variable=self.var).pack(side="left", padx=(0, 12))
        else:
            raise ValueError("Bilinmeyen alan tipi")

        # İlk alandaysak geri butonu pasif
        self.btn_back.configure(state="disabled" if self.idx == 0 else "normal")

    def _collect_current(self):
        f = FIELDS[self.idx]
        key = f["key"]
        raw = self.var.get().strip()

        TEXT_KEYS = { } #5 adet başlık yazınız

        if key == "sizin başlığınız":
            self.answers[key] = normalize_yil(raw)   # 1990 gibi
            return

        if f["type"] == "radio":
            self.answers[key] = raw
            return

        if key in TEXT_KEYS:
            self.answers[key] = turkce_title(raw)
        else:
            self.answers[key] = raw

    def _next(self):
        self._collect_current()
        if self.idx + 1 < len(FIELDS):
            self._show_field(self.idx + 1)
        else:
            self._save_candidate_and_reset()

    def _back(self):
        # Önce mevcut alanı kaydet (ne yazdıysa kaybolmasın)
        self._collect_current()
        if self.idx > 0:
            self._show_field(self.idx - 1)

    def _skip(self):
        key = FIELDS[self.idx]["key"]
        self.answers[key] = ""
        if self.idx + 1 < len(FIELDS):
            self._show_field(self.idx + 1)
        else:
            self._save_candidate_and_reset()

    def _force_finish_candidate(self):
        # Eksik alanlar boş kabul
        self._collect_current()
        for f in FIELDS:
            self.answers.setdefault(f["key"], "")
        self._save_candidate_and_reset()

    def _save_candidate_and_reset(self):
        data = {h: "" for h in BASLIKLAR}
        data["Başvuru Tarihi"] = datetime.now().strftime("%d.%m.%Y")

        for k, v in self.answers.items():
            if k in data:
                data[k] = v

        for col, header in enumerate(BASLIKLAR, start=1):
            self.ws.cell(row=self.current_row, column=col).value = data.get(header, "")

        self.wb.save(DOSYA_ADI)

        messagebox.showinfo("Kaydedildi", f" kaydedildi ✅ (Satır: {self.current_row})\nYeni satıra geçiliyor.")
        self.current_row += 1
        self.answers = {}
        self._show_field(0)

    def _finish(self):
        try:
            self.wb.save(DOSYA_ADI)
        except Exception:
            pass
        self.root.destroy()

    def _on_enter(self, event):
        self._next()

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Excel Kayıt Sistemi")
    root.geometry("820x400")
    root.minsize(720, 340)
    App(root)
    root.mainloop()
